import json
import os
import re
import warnings
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# =========================================================
# 1) 全局路径
# =========================================================
DATA_DIR = "data"
OUTPUT_DIR = "public"

TIER_FILES = {
    "T0": "T0.xlsx",
    "T1": "T1.xlsx",
    "T2": "T2.xlsx",
    "T3": "T3.xlsx",
}

# =========================================================
# 2) 你的仓库（写死）
#   - 退货仓要出现在下拉，但不计算（等你给退货报价数据）
# =========================================================
WAREHOUSES = [
    {"id": "W91730", "label": "SureGo美西库卡蒙格-91730新仓", "zip": "91730", "region": "WEST", "enabled_for_quote": True},
    {"id": "W91752", "label": "SureGo美西米拉罗马-91752仓", "zip": "91752", "region": "WEST", "enabled_for_quote": True},
    {"id": "W60632", "label": "SureGo美中芝加哥-60632仓", "zip": "60632", "region": "CENTRAL", "enabled_for_quote": True},
    {"id": "E08691", "label": "SureGo美东新泽西-08691仓", "zip": "08691", "region": "EAST", "enabled_for_quote": True},
    {"id": "E06801", "label": "SureGo美东贝塞尔-06801仓", "zip": "06801", "region": "EAST", "enabled_for_quote": True},
    {"id": "E11791", "label": "SureGo美东长岛-11791仓", "zip": "11791", "region": "EAST", "enabled_for_quote": True},
    {"id": "E07032", "label": "SureGo美东新泽西-07032仓", "zip": "07032", "region": "EAST", "enabled_for_quote": True},
    {"id": "R63461", "label": "SureGo退货检测-美中密苏里63461退货仓", "zip": "63461", "region": "RETURN", "enabled_for_quote": False},
]

# =========================================================
# 3) 渠道 ↔ 仓库可用映射（按你填写）
#   - 用仓库 region/zip 来判断
#   - 注意：你写了“美西”有多个仓，这里按 region=WEST 覆盖（W91730+W91752）
# =========================================================
CHANNEL_WAREHOUSE_ALLOW = {
    "GOFO-报价": ["WEST", "CENTRAL"],
    "GOFO、UNIUNI-MT-报价": ["WEST", "CENTRAL"],
    "USPS-YSD-报价": ["WEST", "CENTRAL"],
    "FedEx-632-MT-报价": ["WEST", "CENTRAL", "EAST"],
    "FedEx-MT-超大包裹-报价": ["WEST", "CENTRAL", "EAST"],
    "FedEx-ECO-MT报价": ["WEST", "CENTRAL", "EAST"],
    "FedEx-MT-危险品-报价": ["CENTRAL", "EAST"],
    "GOFO大件-MT-报价": ["WEST", "EAST"],
    "XLmiles-报价": ["WEST"],  # 你强调：只有美西可用（主要=91730）；这里按 WEST 放行，但前端会提示“建议91730”
}

# =========================================================
# 4) 费用口径（按你本次填的）
# =========================================================
def money_round(x: float) -> float:
    return float(f"{x:.2f}")

RES_FEES = {
    "FedEx-632-MT-报价": money_round(2.607),
    "FedEx-MT-超大包裹-报价": money_round(2.607),
    "FedEx-MT-危险品-报价": money_round(3.324),
    "GOFO大件-MT-报价": money_round(2.92903225806451),
}

SIG_FEES = {
    "XLmiles-报价": money_round(10.2),
    "FedEx-632-MT-报价": money_round(4.367),
    "FedEx-MT-危险品-报价": money_round(9.708),
    "FedEx-MT-超大包裹-报价": money_round(4.367),
}

# Fuel：哪些渠道额外加燃油
FUEL_CHANNELS = {"FedEx-632-MT-报价", "FedEx-MT-超大包裹-报价", "FedEx-MT-危险品-报价", "GOFO大件-MT-报价"}
# Fuel 85折：仅这两类（你填的）
FUEL_DISCOUNT_85 = {"FedEx-632-MT-报价", "FedEx-MT-超大包裹-报价"}

# =========================================================
# 5) Excel 抽取配置（按你给的固定区块）
#   用 openpyxl 直接读 cell，避免 pandas header 探测失效
# =========================================================
def col_letter_to_index(col: str) -> int:
    from openpyxl.utils.cell import column_index_from_string
    return column_index_from_string(col)

def read_cell(ws, addr: str):
    v = ws[addr].value
    if v is None:
        return ""
    return str(v).strip()

def safe_float(val) -> float:
    try:
        if val is None:
            return 0.0
        s = str(val).strip()
        if s == "" or s.lower() == "nan":
            return 0.0
        s = s.replace("$", "").replace(",", "")
        return float(s)
    except:
        return 0.0

def to_lb_weight(val, unit: str):
    """
    unit: "LB" | "OZ" | "KG"
    """
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return None
    nums = re.findall(r"[\d\.]+", s)
    if not nums:
        return None
    n = float(nums[0])
    if unit == "OZ":
        return n / 16.0
    if unit == "KG":
        return n / 0.453592
    return n

def scan_zone_map(ws, header_row: int, col_start: str, col_end: str):
    """
    扫描 header_row 这一行里 col_start~col_end 的内容，匹配 Zone~n / Zone n / zone~n
    返回：{ "1": "C", "2":"D", ... }
    """
    zmap = {}
    c1 = col_letter_to_index(col_start)
    c2 = col_letter_to_index(col_end)
    for c in range(c1, c2 + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        m = re.search(r"(?:zone|分区)\s*~?\s*(\d+)", s, flags=re.IGNORECASE)
        if m:
            zn = m.group(1)
            if zn not in zmap:
                from openpyxl.utils.cell import get_column_letter
                zmap[zn] = get_column_letter(c)
    return zmap

def extract_table_until_blank(ws, weight_col: str, unit: str, start_row: int, zone_header_row: int, zone_col_start: str, zone_col_end: str):
    """
    通用：按 weight_col 从 start_row 往下读，直到 weight 为空
    zone 列通过 zone_header_row 扫描得到
    返回：prices=[ {w:lb, "2":xx, ...}, ... ], zones=[...]
    """
    zmap = scan_zone_map(ws, zone_header_row, zone_col_start, zone_col_end)
    prices = []
    r = start_row
    while True:
        w_raw = ws[f"{weight_col}{r}"].value
        w_lb = to_lb_weight(w_raw, unit)
        if w_lb is None:
            break
        item = {"w": float(w_lb)}
        for zn, col in zmap.items():
            p = safe_float(ws[f"{col}{r}"].value)
            if p > 0:
                item[zn] = float(p)
        if len(item) > 1:
            prices.append(item)
        r += 1
        if r > 5000:
            break
    prices.sort(key=lambda x: x["w"])
    return list(zmap.keys()), prices

def extract_fedex_dual(ws, res_weight_col, res_start_row, res_zone_header_row, res_zone_start, res_zone_end,
                       com_weight_col, com_start_row, com_zone_header_row, com_zone_start, com_zone_end):
    res_zones, res_prices = extract_table_until_blank(ws, res_weight_col, "LB", res_start_row, res_zone_header_row, res_zone_start, res_zone_end)
    com_zones, com_prices = extract_table_until_blank(ws, com_weight_col, "LB", com_start_row, com_zone_header_row, com_zone_start, com_zone_end)
    return {
        "res": {"zones": res_zones, "prices": res_prices},
        "com": {"zones": com_zones, "prices": com_prices},
    }

def extract_gofo_mixed(ws, zone_header_row: int, zone_col_start: str, zone_col_end: str):
    """
    GOFO-报价：你给的结构是：
      - Zone~1 在 C3
      - OZ weights: A4-A19
      - LB weights: A20 开始
      - KG weights: B4 开始（与 OZ 同行）
    处理策略：
      1) 先读 OZ 行段 A4~A19（unit=OZ）
      2) 再从 A20 往下读 LB（unit=LB）直到空
      3) KG 列（B4）与 OZ 同行容易重复；这里不额外叠加 KG，避免重复档（需要你确认是否为同一档的双单位显示）
    """
    zones = scan_zone_map(ws, zone_header_row, zone_col_start, zone_col_end)
    zkeys = list(zones.keys())

    # 1) OZ: A4~A19 固定
    prices = []
    for r in range(4, 20):
        w_lb = to_lb_weight(ws[f"A{r}"].value, "OZ")
        if w_lb is None:
            continue
        item = {"w": float(w_lb)}
        for zn, col in zones.items():
            p = safe_float(ws[f"{col}{r}"].value)
            if p > 0:
                item[zn] = float(p)
        if len(item) > 1:
            prices.append(item)

    # 2) LB: A20 往下直到空
    r = 20
    while True:
        w_lb = to_lb_weight(ws[f"A{r}"].value, "LB")
        if w_lb is None:
            break
        item = {"w": float(w_lb)}
        for zn, col in zones.items():
            p = safe_float(ws[f"{col}{r}"].value)
            if p > 0:
                item[zn] = float(p)
        if len(item) > 1:
            prices.append(item)
        r += 1
        if r > 5000:
            break

    prices.sort(key=lambda x: x["w"])
    return {"zones": zkeys, "prices": prices}

def extract_usps(ws):
    # USPS-YSD：Zone~1 在 D4，Zone~9 在 L4；weight LB=B4, KG=C4，从第5行开始
    zones = scan_zone_map(ws, 4, "D", "L")
    zkeys = list(zones.keys())
    prices = []
    r = 5
    while True:
        w_lb = to_lb_weight(ws[f"B{r}"].value, "LB")
        if w_lb is None:
            break
        item = {"w": float(w_lb)}
        for zn, col in zones.items():
            p = safe_float(ws[f"{col}{r}"].value)
            if p > 0:
                item[zn] = float(p)
        if len(item) > 1:
            prices.append(item)
        r += 1
        if r > 5000:
            break
    prices.sort(key=lambda x: x["w"])
    return {"zones": zkeys, "prices": prices}

def extract_xlmiles(ws):
    """
    XLmiles：你给的结构（同一张表，Zone列在 D/E/F/G，分别是 Zone~1/2/3/6）
      - AH weights：C4-C8
      - OS weights：C9-C11
      - OM weights：C12-C13
    统一抽取成一个价格表（按 weight 档递增），zone key = 1/2/3/6
    """
    zones = scan_zone_map(ws, 3, "D", "G")
    zkeys = list(zones.keys())
    prices = []
    for r in range(4, 14):
        w_lb = to_lb_weight(ws[f"C{r}"].value, "LB")
        if w_lb is None:
            continue
        item = {"w": float(w_lb)}
        for zn, col in zones.items():
            p = safe_float(ws[f"{col}{r}"].value)
            if p > 0:
                item[zn] = float(p)
        if len(item) > 1:
            prices.append(item)
    prices.sort(key=lambda x: x["w"])
    return {"zones": zkeys, "prices": prices}

def extract_das_amounts(ws):
    """
    你要求：DAS 金额从 G181~G186 自动抽取
    同时把项目名也抽出来（默认 I181~I186，如果为空就用 row index 兜底）
    """
    items = []
    for r in range(181, 187):
        name = ws[f"I{r}"].value
        name = str(name).strip() if name is not None else f"ROW_{r}"
        amt = safe_float(ws[f"G{r}"].value)
        if amt > 0:
            items.append({"name": name, "amount": float(amt)})
    return items

# =========================================================
# 6) 邮编库：仍用 GOFO 邮编区（你没让改）
# =========================================================
US_STATES_CN = {
    "AL": "阿拉巴马", "AK": "阿拉斯加", "AZ": "亚利桑那", "AR": "阿肯色", "CA": "加利福尼亚",
    "CO": "科罗拉多", "CT": "康涅狄格", "DE": "特拉华", "FL": "佛罗里达", "GA": "佐治亚",
    "HI": "夏威夷", "ID": "爱达荷", "IL": "伊利诺伊", "IN": "印第安纳", "IA": "爱荷华",
    "KS": "堪萨斯", "KY": "肯塔基", "LA": "路易斯安那", "ME": "缅因", "MD": "马里兰",
    "MA": "马萨诸塞", "MI": "密歇根", "MN": "明尼苏达", "MS": "密西西比", "MO": "密苏里",
    "MT": "蒙大拿", "NE": "内布拉斯加", "NV": "内华达", "NH": "新罕布什尔", "NJ": "新泽西",
    "NM": "新墨西哥", "NY": "纽约", "NC": "北卡罗来纳", "ND": "北达科他", "OH": "俄亥俄",
    "OK": "俄克拉荷马", "OR": "俄勒冈", "PA": "宾夕法尼亚", "RI": "罗德岛", "SC": "南卡罗来纳",
    "SD": "南达科他", "TN": "田纳西", "TX": "德克萨斯", "UT": "犹他", "VT": "佛蒙特",
    "VA": "弗吉尼亚", "WA": "华盛顿", "WV": "西弗吉尼亚", "WI": "威斯康星", "WY": "怀俄明",
    "DC": "华盛顿特区",
}

# 这里保持你旧 ZIP_COL_MAP 的口径（如果你 GOFO 邮编表有变再调）
ZIP_COL_MAP = {
    "GOFO-报价": 5,
    "GOFO、UNIUNI-MT-报价": 6,
    "UNIUNI-MT-报价": 6,
    "USPS-YSD-报价": 8,
    "FedEx-ECO-MT报价": 9,
    "XLmiles-报价": 10,
    "GOFO大件-MT-报价": 11,
    "FedEx-632-MT-报价": 12,
    "FedEx-MT-超大包裹-报价": 12,
}

def get_sheet_by_exact_or_contains(wb, sheet_name_or_keywords):
    """
    sheet_name_or_keywords: str 或 [keyword1, keyword2...]
    - 若传 str：优先精准匹配，否则 contains
    - 若传 list：contains all keywords
    """
    if isinstance(sheet_name_or_keywords, str):
        # exact
        if sheet_name_or_keywords in wb.sheetnames:
            return wb[sheet_name_or_keywords]
        # contains
        key = sheet_name_or_keywords.replace(" ", "").upper()
        for sn in wb.sheetnames:
            if key in sn.replace(" ", "").upper():
                return wb[sn]
        return None

    keys = [k.replace(" ", "").upper() for k in sheet_name_or_keywords]
    for sn in wb.sheetnames:
        x = sn.replace(" ", "").upper()
        if all(k in x for k in keys):
            return wb[sn]
    return None

def load_zip_db_from_T0():
    """
    从 T0 的 GOFO-报价表里抽 ZIP zone 映射（保持你旧逻辑）
    """
    print("--- 1. 加载邮编库（GOFO独立邮编区） ---")
    path = os.path.join(DATA_DIR, TIER_FILES["T0"])
    if not os.path.exists(path):
        print("❌ 缺少 data/T0.xlsx")
        return {}

    wb = load_workbook(path, data_only=True)
    ws = get_sheet_by_exact_or_contains(wb, ["GOFO", "报价"])
    if ws is None:
        print("❌ 未找到 GOFO-报价 sheet")
        return {}

    # 你旧逻辑：zip 在 B列；从前100行里找第一个5位数字作为起点
    start_row = 1
    for r in range(1, 101):
        v = ws.cell(row=r, column=2).value
        s = str(v).strip() if v is not None else ""
        if s.isdigit() and len(s) == 5:
            start_row = r
            break

    db = {}
    # 列索引：B=2, C=3, D=4, E=5 ...
    for r in range(start_row, ws.max_row + 1):
        z = ws.cell(row=r, column=2).value
        z = str(z).strip().zfill(5) if z is not None else ""
        if not (z.isdigit() and len(z) == 5):
            continue

        region = str(ws.cell(row=r, column=3).value or "").strip()
        st = str(ws.cell(row=r, column=4).value or "").strip().upper()
        city = str(ws.cell(row=r, column=5).value or "").strip()

        zones = {}
        for ch, col_idx in ZIP_COL_MAP.items():
            v = ws.cell(row=r, column=col_idx + 1).value  # 你原表 col map 是按“第几列(从0?)”，这里做 +1 兜底
            # 如果这里不对：你告诉我 zip 表每个渠道 zone 列的真实列号，我直接改这行
            sv = str(v).strip() if v is not None else ""
            zones[ch] = None if sv in ("", "-", "0", "nan", "None") else sv

        db[z] = {
            "s": st,
            "sn": US_STATES_CN.get(st, ""),
            "c": city,
            "r": region,
            "z": zones,
        }

    print(f"✅ 邮编库: {len(db)} 条")
    return db

# =========================================================
# 7) load_tiers：按你固定区块抽取每个 tier 的渠道价格
# =========================================================
def load_tiers_and_das():
    print("\n--- 2. 加载报价表（按固定区块抽取） ---")
    all_tiers = {}
    all_das = {}

    for tier, fname in TIER_FILES.items():
        path = os.path.join(DATA_DIR, fname)
        print(f"处理 {tier}...")
        if not os.path.exists(path):
            print(f"  ❌ 缺少 {path}")
            continue

        wb = load_workbook(path, data_only=True)

        tier_data = {}
        tier_das = {}

        # 1) GOFO-报价
        ws = get_sheet_by_exact_or_contains(wb, "GOFO-报价")
        if ws is not None:
            t = extract_gofo_mixed(ws, zone_header_row=3, zone_col_start="C", zone_col_end="J")
            tier_data["GOFO-报价"] = {"type": "single", "zones": t["zones"], "prices": t["prices"]}
            tier_das["GOFO-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/GOFO-报价: zones={t['zones']}, prices={len(t['prices'])}, das_items={len(tier_das['GOFO-报价'])}")

        # 2) GOFO、UNIUNI-MT-报价（同 sheet 两块表）
        ws = get_sheet_by_exact_or_contains(wb, "GOFO、UNIUNI-MT-报价")
        if ws is not None:
            # GOFO 部分：Zone header 在 C3，weight 从 A3(A4)开始
            gofo_part = extract_table_until_blank(ws, weight_col="A", unit="LB", start_row=3, zone_header_row=3, zone_col_start="C", zone_col_end="J")
            # 但 GOFO-MT 里可能也有 OZ/KB 双单位；这里先用 LB 主表
            gofo_zones, gofo_prices = gofo_part
            # UNIUNI 部分：Zone header 在 N3，weight 在 L3
            uni_zones, uni_prices = extract_table_until_blank(ws, weight_col="L", unit="LB", start_row=3, zone_header_row=3, zone_col_start="N", zone_col_end="U")

            tier_data["GOFO、UNIUNI-MT-报价"] = {
                "type": "combo",
                "gofo": {"zones": gofo_zones, "prices": gofo_prices},
                "uniuni": {"zones": uni_zones, "prices": uni_prices},
            }
            tier_das["GOFO、UNIUNI-MT-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/GOFO、UNIUNI-MT-报价: gofo_prices={len(gofo_prices)}, uni_prices={len(uni_prices)}, das_items={len(tier_das['GOFO、UNIUNI-MT-报价'])}")

        # 3) USPS-YSD-报价
        ws = get_sheet_by_exact_or_contains(wb, "USPS-YSD-报价")
        if ws is not None:
            t = extract_usps(ws)
            tier_data["USPS-YSD-报价"] = {"type": "single", "zones": t["zones"], "prices": t["prices"]}
            tier_das["USPS-YSD-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/USPS-YSD-报价: zones={t['zones']}, prices={len(t['prices'])}, das_items={len(tier_das['USPS-YSD-报价'])}")

        # 4) FedEx-ECO-MT报价（你说即 FedEx-Economy）
        ws = get_sheet_by_exact_or_contains(wb, "FedEx-ECO-MT报价")
        if ws is not None:
            zones, prices = extract_table_until_blank(ws, weight_col="A", unit="LB", start_row=4, zone_header_row=3, zone_col_start="C", zone_col_end="I")
            tier_data["FedEx-ECO-MT报价"] = {"type": "single", "zones": zones, "prices": prices}
            tier_das["FedEx-ECO-MT报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/FedEx-ECO-MT报价: zones={zones}, prices={len(prices)}, das_items={len(tier_das['FedEx-ECO-MT报价'])}")

        # 5) FedEx-632-MT-报价（住宅/商业双表）
        ws = get_sheet_by_exact_or_contains(wb, "FedEx-632-MT-报价")
        if ws is not None:
            t = extract_fedex_dual(
                ws,
                res_weight_col="A", res_start_row=4, res_zone_header_row=3, res_zone_start="C", res_zone_end="I",
                com_weight_col="K", com_start_row=4, com_zone_header_row=3, com_zone_start="M", com_zone_end="S",
            )
            tier_data["FedEx-632-MT-报价"] = {"type": "dual", "res": t["res"], "com": t["com"]}
            tier_das["FedEx-632-MT-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/FedEx-632-MT-报价: res_prices={len(t['res']['prices'])}, com_prices={len(t['com']['prices'])}, das_items={len(tier_das['FedEx-632-MT-报价'])}")

        # 6) FedEx-MT-危险品-报价（住宅/商业双表）
        ws = get_sheet_by_exact_or_contains(wb, "FedEx-MT-危险品-报价")
        if ws is not None:
            t = extract_fedex_dual(
                ws,
                res_weight_col="A", res_start_row=4, res_zone_header_row=3, res_zone_start="C", res_zone_end="I",
                com_weight_col="K", com_start_row=4, com_zone_header_row=3, com_zone_start="M", com_zone_end="S",
            )
            tier_data["FedEx-MT-危险品-报价"] = {"type": "dual", "res": t["res"], "com": t["com"]}
            tier_das["FedEx-MT-危险品-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/FedEx-MT-危险品-报价: res_prices={len(t['res']['prices'])}, com_prices={len(t['com']['prices'])}, das_items={len(tier_das['FedEx-MT-危险品-报价'])}")

        # 7) GOFO大件-MT-报价（住宅/商业双表）
        ws = get_sheet_by_exact_or_contains(wb, "GOFO大件-MT-报价")
        if ws is not None:
            t = extract_fedex_dual(
                ws,
                res_weight_col="A", res_start_row=4, res_zone_header_row=3, res_zone_start="C", res_zone_end="I",
                com_weight_col="K", com_start_row=4, com_zone_header_row=3, com_zone_start="M", com_zone_end="S",
            )
            tier_data["GOFO大件-MT-报价"] = {"type": "dual", "res": t["res"], "com": t["com"]}
            tier_das["GOFO大件-MT-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/GOFO大件-MT-报价: res_prices={len(t['res']['prices'])}, com_prices={len(t['com']['prices'])}, das_items={len(tier_das['GOFO大件-MT-报价'])}")

        # 8) FedEx-MT-超大包裹-报价（住宅/商业双表）
        ws = get_sheet_by_exact_or_contains(wb, "FedEx-MT-超大包裹-报价")
        if ws is not None:
            t = extract_fedex_dual(
                ws,
                res_weight_col="A", res_start_row=4, res_zone_header_row=3, res_zone_start="C", res_zone_end="I",
                com_weight_col="K", com_start_row=4, com_zone_header_row=3, com_zone_start="M", com_zone_end="S",
            )
            tier_data["FedEx-MT-超大包裹-报价"] = {"type": "dual", "res": t["res"], "com": t["com"]}
            tier_das["FedEx-MT-超大包裹-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/FedEx-MT-超大包裹-报价: res_prices={len(t['res']['prices'])}, com_prices={len(t['com']['prices'])}, das_items={len(tier_das['FedEx-MT-超大包裹-报价'])}")

        # 9) XLmiles-报价
        ws = get_sheet_by_exact_or_contains(wb, "XLmiles-报价")
        if ws is not None:
            t = extract_xlmiles(ws)
            tier_data["XLmiles-报价"] = {"type": "single", "zones": t["zones"], "prices": t["prices"]}
            tier_das["XLmiles-报价"] = extract_das_amounts(ws)
            print(f"  > {tier}/XLmiles-报价: zones={t['zones']}, prices={len(t['prices'])}, das_items={len(tier_das['XLmiles-报价'])}")

        all_tiers[tier] = tier_data
        all_das[tier] = tier_das

    return all_tiers, all_das

# =========================================================
# 8) HTML 模板（内嵌 JSON）
# =========================================================
HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>业务员报价助手</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    :root { --header-bg:#000; }
    body { font-family: 'Segoe UI','Microsoft YaHei',sans-serif; background:#f4f6f9; }
    header { background:var(--header-bg); color:#fff; padding:14px 0; }
    .card { border:none; border-radius:10px; box-shadow:0 2px 10px rgba(0,0,0,.06); }
    .card-header{ background:#212529; color:#fff; font-weight:700; }
    .price-text{ font-weight:800; font-size:1.08rem; color:#0d6efd; }
    .small-muted{ color:#6c757d; font-size:.86rem; }
    .mono{ font-family: ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; }
    .result-table th{ background:#212529; color:#fff; text-align:center; font-size:.86rem; }
    .result-table td{ text-align:center; vertical-align:middle; }
  </style>
</head>
<body>
<header>
  <div class="container d-flex justify-content-between align-items-center">
    <div>
      <div class="fw-bold">📦 业务员报价助手</div>
      <div class="opacity-75 small">T0-T3 | 价格更新版</div>
    </div>
    <div class="small">Update: <span id="updateDate"></span></div>
  </div>
</header>

<div class="container my-4">
  <div class="row g-4">
    <div class="col-lg-4">
      <div class="card">
        <div class="card-header">基础信息</div>
        <div class="card-body">
          <div class="mb-3">
            <label class="form-label fw-bold">发货仓库</label>
            <select class="form-select" id="warehouse"></select>
            <div class="small-muted mt-1">退货仓展示但暂不报价（等你补数据）。</div>
          </div>

          <div class="mb-3">
            <label class="form-label fw-bold">客户等级</label>
            <div class="btn-group w-100" role="group">
              <input type="radio" class="btn-check" name="tier" id="t0" value="T0"><label class="btn btn-outline-secondary" for="t0">T0</label>
              <input type="radio" class="btn-check" name="tier" id="t1" value="T1"><label class="btn btn-outline-secondary" for="t1">T1</label>
              <input type="radio" class="btn-check" name="tier" id="t2" value="T2"><label class="btn btn-outline-secondary" for="t2">T2</label>
              <input type="radio" class="btn-check" name="tier" id="t3" value="T3" checked><label class="btn btn-outline-secondary" for="t3">T3</label>
            </div>
          </div>

          <div class="mb-3">
            <label class="form-label fw-bold">目的地邮编</label>
            <div class="input-group">
              <input class="form-control" id="zipCode" placeholder="5位Zip">
              <button class="btn btn-dark" id="btnLookup" type="button">查询</button>
            </div>
            <div class="small mt-1">
              <span id="locInfo" class="text-muted">请输入邮编查询…</span><br/>
              <span id="zoneInfo" class="text-muted"></span>
            </div>
          </div>

          <div class="row g-2 mb-3">
            <div class="col-7">
              <label class="form-label fw-bold">地址类型</label>
              <select class="form-select" id="addressType">
                <option value="res">住宅 Residential</option>
                <option value="com">商业 Commercial</option>
              </select>
            </div>
            <div class="col-5 pt-4">
              <div class="form-check form-switch">
                <input class="form-check-input" type="checkbox" id="sigToggle">
                <label class="form-check-label fw-bold" for="sigToggle">签名签收</label>
              </div>
            </div>
          </div>

          <div class="bg-light p-2 rounded border mb-3">
            <div class="fw-bold small mb-2 border-bottom">⛽ 燃油费率 (Fuel)</div>
            <div class="small text-danger fw-bold mb-2">仅：FedEx-632 / FedEx-MT-超大包裹 / FedEx-危险品 / GOFO大件</div>
            <div class="row g-2">
              <div class="col-12">
                <label class="form-label small">FedEx Fuel (%)</label>
                <input type="number" class="form-control form-control-sm" id="fedexFuel" value="16.0">
                <div class="small-muted mt-1">FedEx-632 / 超大包裹：燃油按 85 折；危险品不打折。</div>
              </div>
              <div class="col-12 mt-2">
                <label class="form-label small">GOFO大件 Fuel (%)</label>
                <input type="number" class="form-control form-control-sm" id="gofoFuel" value="15.0">
              </div>
            </div>
          </div>

          <hr/>

          <div class="mb-3">
            <label class="form-label fw-bold">包裹规格</label>
            <div class="row g-2">
              <div class="col-4"><input class="form-control form-control-sm" id="length" placeholder="长(in)"></div>
              <div class="col-4"><input class="form-control form-control-sm" id="width" placeholder="宽(in)"></div>
              <div class="col-4"><input class="form-control form-control-sm" id="height" placeholder="高(in)"></div>
              <div class="col-8"><input class="form-control form-control-sm" id="weight" placeholder="重量(lb)"></div>
              <div class="col-4">
                <select class="form-select form-select-sm" id="weightUnit">
                  <option value="lb">lb</option>
                  <option value="oz">oz</option>
                  <option value="kg">kg</option>
                  <option value="g">g</option>
                </select>
              </div>
            </div>
            <div class="small-muted mt-1">尺寸单位固定按英寸；重量支持 lb/oz/kg/g。</div>
          </div>

          <button class="btn btn-primary w-100 fw-bold" id="btnCalc" type="button">开始计算</button>
        </div>
      </div>
    </div>

    <div class="col-lg-8">
      <div class="card">
        <div class="card-header d-flex justify-content-between align-items-center">
          <span>📊 报价结果</span>
          <span class="badge bg-warning text-dark" id="tierBadge"></span>
        </div>
        <div class="card-body">
          <div class="alert alert-info py-2 small" id="pkgSummary">请先输入数据…</div>

          <div class="border rounded p-2 mb-3">
            <div class="fw-bold">说明（关键口径）</div>
            <div class="small mt-1" style="line-height:1.35">
              1) <b>只显示当前仓库可用渠道</b>。退货仓暂不报价。<br/>
              2) <b>FedEx Zone</b>：按「发货仓邮编 + 目的邮编」计算（不依赖 GOFO 邮编区）。<br/>
              3) <b>住宅费</b>：按渠道固定金额叠加（见明细）。<br/>
              4) <b>签名费</b>：开关控制叠加；仅对指定渠道生效（见明细）。<br/>
              5) <b>燃油</b>：仅对指定渠道额外计算；其中 FedEx-632/超大包裹燃油按 85 折。<br/>
              6) <b>XLmiles</b>：一口价包含保价/预约及签收证明等服务；“第二件起半价/一票多件分摊公式”仅一票多件适用，本工具当前按单件报价，不自动套用分摊公式。<br/>
              <div class="small text-muted mt-1 mono">DAS 金额已从 Excel G181~G186 注入 JSON（不自动计入，等你给 ZIP 判定口径）。</div>
            </div>
          </div>

          <div class="table-responsive">
            <table class="table table-bordered table-hover result-table">
              <thead>
                <tr>
                  <th width="22%">渠道</th>
                  <th width="18%">仓库</th>
                  <th width="10%">Zone</th>
                  <th width="10%">计费重(lb)</th>
                  <th width="12%">基础运费</th>
                  <th width="18%">附加费明细</th>
                  <th width="10%">总费用</th>
                </tr>
              </thead>
              <tbody id="resBody"></tbody>
            </table>
          </div>

        </div>
      </div>
    </div>

  </div>
</div>

<script>
  let DATA = __JSON_DATA__;

  document.getElementById("updateDate").innerText = new Date().toLocaleDateString();

  // 渲染仓库下拉
  (function initWarehouses(){
    const sel = document.getElementById("warehouse");
    DATA.warehouses.forEach(w=>{
      const opt = document.createElement("option");
      opt.value = w.id;
      opt.textContent = w.label;
      sel.appendChild(opt);
    });
    sel.value = DATA.warehouses[0].id;
  })();

  // 自动计算监听
  document.querySelectorAll('input[name="tier"]').forEach(r => r.addEventListener('change', ()=>document.getElementById('btnCalc').click()));
  document.getElementById('warehouse').addEventListener('change', ()=>document.getElementById('btnCalc').click());
  document.getElementById('addressType').addEventListener('change', ()=>document.getElementById('btnCalc').click());
  document.getElementById('sigToggle').addEventListener('change', ()=>document.getElementById('btnCalc').click());

  // ====== 工具函数 ======
  function money(x){ return Number.parseFloat(x||0).toFixed(2); }
  function stdWeight(w, unit){
    let v = parseFloat(w||0);
    if(!v || v<0) return 0;
    if(unit==="oz") return v/16;
    if(unit==="kg") return v/0.453592;
    if(unit==="g") return v/453.592;
    return v;
  }
  function pkgSummary(pkg){
    const dims = [pkg.L,pkg.W,pkg.H].sort((a,b)=>b-a);
    const G = dims[0] + 2*(dims[1]+dims[2]);
    return {dims, G};
  }

  // ====== FedEx Zone 计算：按仓邮编前三位(origin3) + 目的邮编前三位(dest3) ======
  function calcFedexZone(destZip, originZip){
    if(!destZip || destZip.length<3) return null;
    const p = parseInt(destZip.substring(0,3), 10);
    const o = parseInt(originZip.substring(0,3), 10);

    // AK/HI/PR/VI/Guam 等粗暴兜底（你要精确表我再按表做）
    if ((p >= 967 && p <= 969) || (p >= 995 && p <= 999) || destZip.startsWith("00")) return 9;

    // origin 分三类：917 / 606 / 0xx(东部)
    const origin3 = String(o).padStart(3,"0");
    if(origin3==="917"){
      if (p >= 900 && p <= 935) return 2;
      if (p >= 936 && p <= 961) return 3;
      if (p >= 890 && p <= 898) return 3;
      if (p >= 970 && p <= 994) return 4;
      if (p >= 840 && p <= 884) return 4;
      if (p >= 500 && p <= 799) return 6;
      if (p >= 0 && p <= 499) return 8;
      return 8;
    }
    if(origin3==="606"){
      if (p >= 600 && p <= 629) return 2;
      if (p >= 460 && p <= 569) return 3;
      if (p >= 400 && p <= 459) return 4;
      if (p >= 700 && p <= 799) return 4;
      if (p >= 200 && p <= 399) return 5;
      if (p >= 800 && p <= 899) return 6;
      if (p >= 0 && p <= 199) return 7;
      if (p >= 900 && p <= 966) return 8;
      return 8;
    }
    // EAST：068/070/086/117 等按“东部”近似
    if (p >= 70 && p <= 89) return 2;
    if (p >= 0 && p <= 69) return 3;
    if (p >= 150 && p <= 199) return 3;
    if (p >= 200 && p <= 299) return 4;
    if (p >= 400 && p <= 599) return 5;
    if (p >= 600 && p <= 799) return 7;
    if (p >= 800 && p <= 966) return 8;
    return 8;
  }

  function getWarehouseById(id){
    return DATA.warehouses.find(w=>w.id===id);
  }

  // ====== GOFO 邮编库查州/城 + 其它渠道 zone（USPS/GOFO） ======
  let CUR_ZONES = {};
  let LAST_LOC = null;

  async function lookupZip(zip){
    CUR_ZONES = {};
    LAST_LOC = null;
    const loc = document.getElementById("locInfo");
    const zi = document.getElementById("zoneInfo");

    if(DATA.zip_db && DATA.zip_db[zip]){
      const i = DATA.zip_db[zip];
      loc.innerHTML = `✅ ${i.sn} ${i.s} - ${i.c} [${i.r}]`;
      CUR_ZONES = i.z || {};
      LAST_LOC = {state:i.s, city:i.c};
    }else{
      loc.innerHTML = `⚠️ GOFO邮编库无该邮编`;
    }

    const wh = getWarehouseById(document.getElementById("warehouse").value);
    if(wh && zip && zip.length===5){
      const z = calcFedexZone(zip, wh.zip);
      zi.innerHTML = z ? `FedEx Zone(按仓库计算): <b>Zone ${z}</b>` : ``;
    }else{
      zi.innerHTML = ``;
    }
  }

  document.getElementById("btnLookup").onclick = async ()=>{
    const zip = (document.getElementById("zipCode").value||"").trim();
    if(zip.length!==5){ alert("请输入5位邮编"); return; }
    await lookupZip(zip);
  };

  function channelAllowedInWarehouse(channelName, wh){
    const allow = DATA.channel_allow[channelName] || [];
    return allow.includes(wh.region);
  }

  function pickPriceRow(prices, billW){
    if(!prices || prices.length===0) return null;
    for(const r of prices){
      if((r.w||0) >= billW - 1e-9) return r;
    }
    return null;
  }

  function getZoneKeyForPrice(channelName, zoneVal){
    // 你明确：FedEx 多数从 Zone~2 开始；若 zone=1，使用 zone=2
    if(zoneVal===null || zoneVal===undefined) return null;
    let z = parseInt(zoneVal,10);
    if(Number.isNaN(z)) return null;

    // XLmiles 表只有 1/2/3/6：如果 FedEx zone>=4，映射到 6
    if(channelName.includes("XLmiles")){
      if(z===1 || z===2) return "1";
      if(z===3) return "3";
      if(z>=4) return "6";
      return null;
    }

    if(z===1) z = 2;
    return String(z);
  }

  function isFedexZoneChannel(ch){
    return ch.includes("FedEx") || ch.includes("GOFO大件") || ch.includes("XLmiles");
  }

  function getResFee(ch){
    return DATA.fees.res[ch] || 0;
  }
  function getSigFee(ch){
    return DATA.fees.sig[ch] || 0;
  }
  function hasFuel(ch){
    return DATA.fuel.channels.includes(ch);
  }
  function fuelRateForChannel(ch, baseRate){
    // baseRate 已是小数
    if(DATA.fuel.discount85.includes(ch)) return baseRate * 0.85;
    return baseRate;
  }

  document.getElementById("btnCalc").onclick = async ()=>{
    const zip = (document.getElementById("zipCode").value||"").trim();
    if(zip.length===5 && (!LAST_LOC && Object.keys(CUR_ZONES||{}).length===0)){
      await lookupZip(zip);
    }

    const tier = document.querySelector('input[name="tier"]:checked').value;
    document.getElementById("tierBadge").innerText = tier;

    const wh = getWarehouseById(document.getElementById("warehouse").value);
    const isRes = document.getElementById("addressType").value==="res";
    const sigOn = document.getElementById("sigToggle").checked;

    // 退货仓：直接提示不报价
    const tbody = document.getElementById("resBody");
    tbody.innerHTML = "";
    if(!wh.enabled_for_quote){
      document.getElementById("pkgSummary").innerHTML = `<b>提示：</b> 当前选择的是退货仓（暂未接入退货报价数据）。`;
      tbody.innerHTML = `<tr><td colspan="7" class="text-muted">退货仓报价未接入</td></tr>`;
      return;
    }

    const pkg = {
      L: parseFloat(document.getElementById("length").value||0),
      W: parseFloat(document.getElementById("width").value||0),
      H: parseFloat(document.getElementById("height").value||0),
      Wt: stdWeight(document.getElementById("weight").value, document.getElementById("weightUnit").value)
    };
    const s = pkgSummary(pkg);
    document.getElementById("pkgSummary").innerHTML =
      `<b>基准:</b> ${s.dims[0].toFixed(1)}"×${s.dims[1].toFixed(1)}"×${s.dims[2].toFixed(1)}" | 实重:${pkg.Wt.toFixed(2)}lb | 围长:${s.G.toFixed(1)}"`;

    const fedexFuel = parseFloat(document.getElementById("fedexFuel").value||0)/100.0;
    const gofoFuel = parseFloat(document.getElementById("gofoFuel").value||0)/100.0;

    const tierData = (DATA.tiers && DATA.tiers[tier]) ? DATA.tiers[tier] : null;
    if(!tierData){
      tbody.innerHTML = `<tr><td colspan="7" class="text-danger">缺少 ${tier} 数据</td></tr>`;
      return;
    }

    // FedEx zone（统一使用）
    const fedexZone = (zip.length===5) ? calcFedexZone(zip, wh.zip) : null;

    // 迭代所有渠道，但只显示“仓可用”
    Object.keys(tierData).forEach(ch=>{
      if(!channelAllowedInWarehouse(ch, wh)) return;

      // zoneVal：FedEx类用计算；USPS/GOFO 用 GOFO 邮编库
      let zoneVal = null;
      if(isFedexZoneChannel(ch)){
        zoneVal = fedexZone;
      }else{
        const z = (CUR_ZONES && CUR_ZONES[ch]) ? CUR_ZONES[ch] : null;
        zoneVal = z ? parseInt(z,10) : null;
      }

      const zoneKey = getZoneKeyForPrice(ch, zoneVal);

      // 计费重：先按实重，>=1lb 向上取整（保持你之前口径）
      let billW = pkg.Wt;
      if(billW>1) billW = Math.ceil(billW);

      let base = 0;
      let details = [];
      let status = "OK";

      // 取价格
      const chObj = tierData[ch];

      if(!zoneKey){
        status = "无Zone";
      }else{
        if(chObj.type==="single"){
          const row = pickPriceRow(chObj.prices, billW);
          if(!row || row[zoneKey]===undefined){
            status = "无报价";
          }else{
            base = Number(row[zoneKey]||0);
          }
        }else if(chObj.type==="dual"){
          const table = isRes ? chObj.res : chObj.com;
          const row = pickPriceRow(table.prices, billW);
          if(!row || row[zoneKey]===undefined){
            status = "无报价";
          }else{
            base = Number(row[zoneKey]||0);
          }
        }else if(chObj.type==="combo"){
          // 你合并的 GOFO+UNIUNI：这里展示为一个合并渠道，但按“两个子表分别算一行”会更清晰。
          // 为避免你客户误读：这里拆两行展示（GOFO-MT 与 UNIUNI-MT）
          // ——直接在前端拆，不改 Excel 结构
          const gofoRow = pickPriceRow(chObj.gofo.prices, billW);
          const uniRow  = pickPriceRow(chObj.uniuni.prices, billW);

          const zoneK = zoneKey; // GOFO/UNIUNI zone 结构与 GOFO邮编库一致
          const gofoBase = (gofoRow && gofoRow[zoneK]!==undefined) ? Number(gofoRow[zoneK]||0) : 0;
          const uniBase  = (uniRow && uniRow[zoneK]!==undefined)  ? Number(uniRow[zoneK]||0)  : 0;

          // GOFO 子行
          if(gofoBase>0){
            let t = gofoBase;
            const d = [];
            // GOFO-MT 是否额外燃油？你没要求对它加燃油（默认表价已含燃油）
            tbody.innerHTML += `<tr>
              <td class="fw-bold text-start text-nowrap">GOFO-MT</td>
              <td class="text-nowrap">${wh.label}</td>
              <td>Zone ${zoneVal}</td>
              <td>${billW.toFixed(2)}</td>
              <td class="fw-bold">${money(gofoBase)}</td>
              <td class="text-start small">${d.join("<br>")||"-"}</td>
              <td class="price-text">$${money(t)}</td>
            </tr>`;
          }
          // UNIUNI 子行
          if(uniBase>0){
            let t = uniBase;
            const d = [];
            tbody.innerHTML += `<tr>
              <td class="fw-bold text-start text-nowrap">UNIUNI-MT</td>
              <td class="text-nowrap">${wh.label}</td>
              <td>Zone ${zoneVal}</td>
              <td>${billW.toFixed(2)}</td>
              <td class="fw-bold">${money(uniBase)}</td>
              <td class="text-start small">${d.join("<br>")||"-"}</td>
              <td class="price-text">$${money(t)}</td>
            </tr>`;
          }
          return;
        }
      }

      // 费用叠加（base>0才叠加）
      let total = base;
      if(base>0){
        // 住宅费
        if(isRes){
          const rf = getResFee(ch);
          if(rf>0){
            details.push(`住宅:$${money(rf)}`);
            total += rf;
          }
        }

        // 签名费
        if(sigOn){
          const sf = getSigFee(ch);
          if(sf>0){
            details.push(`签名:$${money(sf)}`);
            total += sf;
          }
        }

        // 燃油
        if(hasFuel(ch)){
          if(ch.includes("GOFO大件")){
            const f = total * gofoFuel; // GOFO大件：按(基础+附加)乘燃油
            details.push(`燃油:$${money(f)}`);
            total += f;
          }else{
            const rate = fuelRateForChannel(ch, fedexFuel);
            const f = base * rate;      // FedEx：按基础运费乘燃油
            details.push(`燃油:$${money(f)}`);
            total += f;
          }
        }
      }

      tbody.innerHTML += `<tr>
        <td class="fw-bold text-start text-nowrap">${ch}</td>
        <td class="text-nowrap">${wh.label}</td>
        <td>${zoneVal ? ("Zone "+zoneVal) : "-"}</td>
        <td>${billW.toFixed(2)}</td>
        <td class="fw-bold">${money(base)}</td>
        <td class="text-start small">${details.join("<br>")||"-"}</td>
        <td class="price-text">${total>0?("$"+money(total)):"-"}</td>
      </tr>`;
    });
  };
</script>

</body>
</html>
"""

# =========================================================
# 9) 入口：生成 public/index.html
# =========================================================
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    zip_db = load_zip_db_from_T0()
    tiers, das = load_tiers_and_das()

    final = {
        "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "warehouses": WAREHOUSES,
        "channel_allow": CHANNEL_WAREHOUSE_ALLOW,
        "fees": {"res": RES_FEES, "sig": SIG_FEES},
        "fuel": {"channels": sorted(list(FUEL_CHANNELS)), "discount85": sorted(list(FUEL_DISCOUNT_85))},
        "zip_db": zip_db,
        "tiers": tiers,
        "das_amounts": das,  # ✅ 已注入（按 tier/channel -> items）
    }

    js_str = json.dumps(final, ensure_ascii=False)
    html = HTML_TEMPLATE.replace("__JSON_DATA__", js_str)

    out_path = os.path.join(OUTPUT_DIR, "index.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    print("\n--- 3. 生成网页 ---")
    print(f"✅ 已生成: {out_path}")
